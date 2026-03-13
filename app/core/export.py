"""数据导出服务模块

此模块提供数据导出功能，支持 Excel、CSV 等格式。
"""
import io
import csv
import logging
from typing import List, Dict, Any, Optional
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.core.config import config

logger = logging.getLogger(__name__)


class ExportService:
    """数据导出服务类"""
    
    def __init__(self):
        self.export_dir = Path(config.EXPORT_DIR)
        self.export_dir.mkdir(parents=True, exist_ok=True)
    
    def export_to_csv(
        self,
        data: List[Dict[str, Any]],
        filename: str,
        columns: Optional[List[str]] = None
    ) -> str:
        """导出数据到 CSV 文件"""
        try:
            if not data:
                logger.warning("导出的数据为空")
                return ""
            
            if columns:
                fieldnames = columns
            else:
                fieldnames = list(data[0].keys())
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            export_filename = f"{filename}_{timestamp}.csv"
            export_path = self.export_dir / export_filename
            
            with open(export_path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
                writer.writeheader()
                writer.writerows(data)
            
            logger.info(f"CSV 导出成功：{export_path}")
            return str(export_path)
            
        except Exception as e:
            logger.error(f"CSV 导出失败：{str(e)}")
            return ""
    
    def export_to_excel(
        self,
        data: List[Dict[str, Any]],
        filename: str,
        columns: Optional[Dict[str, str]] = None,
        sheet_name: str = "Sheet1"
    ) -> str:
        """导出数据到 Excel 文件"""
        try:
            if not data:
                logger.warning("导出的数据为空")
                return ""
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            export_filename = f"{filename}_{timestamp}.xlsx"
            export_path = self.export_dir / export_filename
            
            df = pd.DataFrame(data)
            
            if columns:
                df = df[[col for col in columns.keys() if col in df.columns]]
                df.columns = [columns.get(col, col) for col in df.columns]
            
            with pd.ExcelWriter(export_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]
                
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="667eea", end_color="764ba2", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                for col in worksheet.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.2
                    worksheet.column_dimensions[column].width = min(adjusted_width, 50)
                
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    for cell in row:
                        cell.border = thin_border
            
            logger.info(f"Excel 导出成功：{export_path}")
            return str(export_path)
            
        except Exception as e:
            logger.error(f"Excel 导出失败：{str(e)}")
            return ""
    
    def export_to_csv_bytes(
        self,
        data: List[Dict[str, Any]],
        columns: Optional[List[str]] = None
    ) -> bytes:
        """导出数据到 CSV 字节流"""
        try:
            if not data:
                return b""
            
            if columns:
                fieldnames = columns
            else:
                fieldnames = list(data[0].keys())
            
            output = io.StringIO()
            writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction='ignore')
            writer.writeheader()
            writer.writerows(data)
            
            return output.getvalue().encode('utf-8-sig')
            
        except Exception as e:
            logger.error(f"CSV 字节流导出失败：{str(e)}")
            return b""
    
    def export_to_excel_bytes(
        self,
        data: List[Dict[str, Any]],
        columns: Optional[Dict[str, str]] = None,
        sheet_name: str = "Sheet1"
    ) -> bytes:
        """导出数据到 Excel 字节流"""
        try:
            if not data:
                return b""
            
            df = pd.DataFrame(data)
            
            if columns:
                df = df[[col for col in columns.keys() if col in df.columns]]
                df.columns = [columns.get(col, col) for col in df.columns]
            
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
            output.seek(0)
            
            return output.getvalue()
            
        except Exception as e:
            logger.error(f"Excel 字节流导出失败：{str(e)}")
            return b""
    
    def cleanup_old_exports(self, days: int = 7) -> int:
        """清理旧的导出文件"""
        try:
            count = 0
            now = datetime.now()
            
            for file_path in self.export_dir.glob("*"):
                if file_path.is_file():
                    file_time = datetime.fromtimestamp(file_path.stat().st_mtime)
                    age = (now - file_time).days
                    
                    if age > days:
                        file_path.unlink()
                        count += 1
            
            logger.info(f"清理了 {count} 个旧的导出文件")
            return count
            
        except Exception as e:
            logger.error(f"清理导出文件失败：{str(e)}")
            return 0


export_service = ExportService()
